#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
請求書入力自動化ツール（Streamlit Web版）
Excel テンプレート（請求書(ひな形）.xlsx）を使って請求書を自動生成します。
st.data_editor を使わず st.text_input ベースで明細を構築し、
Enter キーによる rerun で値が消える問題を回避しています。
"""

import streamlit as st
from openpyxl import load_workbook
from datetime import date
import os
import re
import io
import math

# ---------- 定数 ----------
_HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_HERE, "請求書(ひな形）.xlsx")
MAX_ITEMS = 16
TAX_OPTIONS = ["10%", "8%", "非課税"]

# ---------- ページ設定 ----------
st.set_page_config(page_title="請求書入力ツール", layout="wide")

# ---------- CSS ----------
st.markdown("""
<style>
/* 明細テーブルのラベル行 */
div[data-testid="stHorizontalBlock"] .item-header {
    background: #cce4f0; font-weight: bold; text-align: center;
    padding: 6px 4px; border-radius: 4px; font-size: 0.85rem;
}
/* 入力欄の上マージンを詰める */
div[data-testid="stVerticalBlock"] > div { gap: 0.25rem; }
/* 小計列の表示 */
.subtotal-cell {
    background: #f0f0f0; border: 1px solid #ddd; border-radius: 4px;
    padding: 8px 6px; text-align: right; font-size: 0.9rem;
    min-height: 38px; line-height: 22px;
}
/* 合計エリア */
.total-box {
    background: #fafafa; border: 1px solid #ddd; border-radius: 4px;
    padding: 6px 10px; text-align: right; font-size: 1rem;
}
.total-box-grand {
    background: #fafafa; border: 2px solid #333; border-radius: 4px;
    padding: 6px 10px; text-align: right; font-size: 1.1rem; font-weight: bold;
}
</style>
""", unsafe_allow_html=True)


# ================================================================
#  セッションステート初期化
# ================================================================
def _init_state():
    today = date.today()
    defaults = {
        "client": "",
        "invoice_no": "",
        "inv_date": f"{today.year}/{today.month:02d}/{today.day:02d}",
        "deadline": "",
        "subject": "",
        "remarks": "",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    if "items" not in st.session_state:
        st.session_state["items"] = []
        for _ in range(MAX_ITEMS):
            st.session_state["items"].append({
                "tax": "10%",
                "hinban": "",
                "hinmei": "",
                "qty": "",
                "unit": "",
                "price": "",
            })


_init_state()


# ================================================================
#  計算ヘルパー
# ================================================================
def _safe_float(s: str) -> float:
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _row_subtotal(row: dict) -> float:
    return _safe_float(row["qty"]) * _safe_float(row["price"])


def _calc_totals():
    subtotal = base8 = base10 = exempt = 0.0
    for row in st.session_state["items"]:
        v = _row_subtotal(row)
        if v:
            t = row["tax"]
            subtotal += v
            if t == "8%":
                base8 += v
            elif t == "非課税":
                exempt += v
            else:
                base10 += v
    t8 = base8 * 0.08
    t10 = base10 * 0.10
    total = subtotal + t8 + t10
    return {
        "subtotal": subtotal, "base10": base10, "tax10": t10,
        "base8": base8, "tax8": t8, "exempt": exempt, "total": total,
    }


# ================================================================
#  Excel 書き込み
# ================================================================
def _parse_date(s: str):
    m = re.match(r"(\d{4})[/\-年](\d{1,2})[/\-月](\d{1,2})", s.strip())
    if m:
        return int(m.group(1)), int(m.group(2)), int(m.group(3))
    t = date.today()
    return t.year, t.month, t.day


def _write_excel() -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    ss = st.session_state

    # 請求日
    y, mo, d = _parse_date(ss["inv_date"])
    ws["H4"] = f"請求日　　{y}年{mo:02d}月{d:02d}日"

    # 請求番号
    inv = ss["invoice_no"].strip()
    ws["H5"] = f"請求番号　{inv}" if inv else "請求番号"

    # 請求先
    ws["C8"] = f"　{ss['client'].strip()}　御中"

    # 件名
    ws["D14"] = ss["subject"].strip()

    # 支払い期限
    ws["E17"] = ss["deadline"].strip()

    # 明細行 (行 22〜37)
    for i, row in enumerate(ss["items"]):
        r = 22 + i
        hinmei = row["hinmei"].strip()
        qty_s = row["qty"].strip()
        price_s = row["price"].strip()

        if not (hinmei or qty_s or price_s):
            for c in range(2, 9):
                ws.cell(row=r, column=c).value = None
            continue

        tax = row["tax"]
        code = "8%" if tax == "8%" else ("非" if tax == "非課税" else "")

        ws.cell(row=r, column=2).value = code
        ws.cell(row=r, column=3).value = row["hinban"].strip() or None
        ws.cell(row=r, column=4).value = hinmei or None

        try:
            ws.cell(row=r, column=5).value = float(qty_s) if qty_s else None
        except ValueError:
            ws.cell(row=r, column=5).value = None

        ws.cell(row=r, column=6).value = row["unit"].strip() or None

        try:
            ws.cell(row=r, column=7).value = float(price_s) if price_s else None
        except ValueError:
            ws.cell(row=r, column=7).value = None

        ws.cell(row=r, column=8).value = f"=E{r}*G{r}"

    # 税計算フォーミュラ
    ws["E40"] = '=SUMIF(B22:B37,"8%",H22:H37)'
    ws["G40"] = "=H38-E40-H40"

    # 備考
    remarks = ss["remarks"].strip()
    if remarks:
        ws["B39"] = remarks

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ================================================================
#  UI
# ================================================================
st.title("請求書入力ツール")

# ---------- 基本情報 ----------
with st.container():
    st.subheader("基本情報")
    c1, c2 = st.columns(2)
    with c1:
        st.session_state["client"] = st.text_input(
            "請求先（御中）*", value=st.session_state["client"], key="w_client")
        st.session_state["inv_date"] = st.text_input(
            "請求日 (YYYY/MM/DD)", value=st.session_state["inv_date"], key="w_date")
        st.session_state["subject"] = st.text_input(
            "件名", value=st.session_state["subject"], key="w_subject")
    with c2:
        st.session_state["invoice_no"] = st.text_input(
            "請求番号", value=st.session_state["invoice_no"], key="w_invno")
        st.session_state["deadline"] = st.text_input(
            "支払い期限 (例: 2024年3月31日)",
            value=st.session_state["deadline"], key="w_deadline")

# ---------- 明細テーブル ----------
st.subheader("明細")
st.caption("※ 税区分: 10% = 標準税率（無記入と同じ）／ 8% = 軽減税率 ／ 非課税")

# ヘッダー行
hcols = st.columns([1.0, 1.2, 3.0, 1.0, 0.8, 1.4, 1.4])
headers = ["税区分", "品番", "品名・品番", "数量", "単位", "単価", "小計"]
for col, label in zip(hcols, headers):
    col.markdown(f'<div class="item-header">{label}</div>', unsafe_allow_html=True)

# 各明細行
for i in range(MAX_ITEMS):
    row = st.session_state["items"][i]
    cols = st.columns([1.0, 1.2, 3.0, 1.0, 0.8, 1.4, 1.4])

    with cols[0]:
        row["tax"] = st.selectbox(
            f"税{i}", TAX_OPTIONS,
            index=TAX_OPTIONS.index(row["tax"]),
            key=f"tax_{i}", label_visibility="collapsed")
    with cols[1]:
        row["hinban"] = st.text_input(
            f"品番{i}", value=row["hinban"],
            key=f"hinban_{i}", label_visibility="collapsed")
    with cols[2]:
        row["hinmei"] = st.text_input(
            f"品名{i}", value=row["hinmei"],
            key=f"hinmei_{i}", label_visibility="collapsed")
    with cols[3]:
        row["qty"] = st.text_input(
            f"数量{i}", value=row["qty"],
            key=f"qty_{i}", label_visibility="collapsed")
    with cols[4]:
        row["unit"] = st.text_input(
            f"単位{i}", value=row["unit"],
            key=f"unit_{i}", label_visibility="collapsed")
    with cols[5]:
        row["price"] = st.text_input(
            f"単価{i}", value=row["price"],
            key=f"price_{i}", label_visibility="collapsed")
    with cols[6]:
        sub = _row_subtotal(row)
        display = f"¥{sub:,.0f}" if sub else ""
        cols[6].markdown(
            f'<div class="subtotal-cell">{display}</div>',
            unsafe_allow_html=True)

# ---------- 合計 ----------
st.subheader("合計（自動計算）")
totals = _calc_totals()

tc1, tc2, tc3 = st.columns(3)
with tc1:
    st.markdown("**小計**")
    st.markdown(f'<div class="total-box">¥{totals["subtotal"]:,.0f}</div>',
                unsafe_allow_html=True)
with tc2:
    st.markdown("**10%対象**")
    st.markdown(f'<div class="total-box">¥{totals["base10"]:,.0f}</div>',
                unsafe_allow_html=True)
    st.markdown("**8%対象**")
    st.markdown(f'<div class="total-box">¥{totals["base8"]:,.0f}</div>',
                unsafe_allow_html=True)
    st.markdown("**非課税**")
    st.markdown(f'<div class="total-box">¥{totals["exempt"]:,.0f}</div>',
                unsafe_allow_html=True)
with tc3:
    st.markdown("**消費税(10%)**")
    st.markdown(f'<div class="total-box">¥{totals["tax10"]:,.0f}</div>',
                unsafe_allow_html=True)
    st.markdown("**消費税(8%)**")
    st.markdown(f'<div class="total-box">¥{totals["tax8"]:,.0f}</div>',
                unsafe_allow_html=True)

st.divider()
st.markdown("### 御請求金額（税込）")
st.markdown(f'<div class="total-box-grand">¥{totals["total"]:,.0f}</div>',
            unsafe_allow_html=True)

# ---------- 備考 ----------
st.subheader("備考・支払い条件")
st.session_state["remarks"] = st.text_area(
    "備考", value=st.session_state["remarks"], key="w_remarks",
    placeholder="例: 翌月末日払い、銀行振込にてお願いいたします。", height=120)

# ---------- ボタン ----------
st.divider()
btn_cols = st.columns([1, 1, 4])

with btn_cols[0]:
    if st.button("フォームをクリア", use_container_width=True):
        today = date.today()
        st.session_state["client"] = ""
        st.session_state["invoice_no"] = ""
        st.session_state["inv_date"] = f"{today.year}/{today.month:02d}/{today.day:02d}"
        st.session_state["deadline"] = ""
        st.session_state["subject"] = ""
        st.session_state["remarks"] = ""
        for row in st.session_state["items"]:
            row["tax"] = "10%"
            for k in ("hinban", "hinmei", "qty", "unit", "price"):
                row[k] = ""
        st.rerun()

with btn_cols[1]:
    # バリデーション
    client_ok = bool(st.session_state["client"].strip())
    has_item = any(r["hinmei"].strip() for r in st.session_state["items"])
    can_export = client_ok and has_item

    if not can_export:
        st.button("Excel 出力", use_container_width=True, disabled=True)
        if not client_ok:
            st.caption("⚠ 請求先を入力してください")
        elif not has_item:
            st.caption("⚠ 明細を1行以上入力してください")
    else:
        today_str = date.today().strftime("%Y%m%d")
        filename = f"請求書_{today_str}_{st.session_state['client'].strip()}.xlsx"
        excel_bytes = _write_excel()
        st.download_button(
            label="Excel 出力",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
